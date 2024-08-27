import tkinter
from tkinter import ttk
from tkinter import messagebox
import os
import tkinter.messagebox
import openpyxl
import pandas as pd


formField_list = []
row_number = []
record_list = []
record_ids_dic = {}
class_type_list = ['','nusery', 'class', 'JSS', 'SSS']
current_sheet = []
cwd = os.getcwd()
filepath = cwd+"\\data.xlsx" # make sure to update the path syntax for other OS

# This class creates an external window
class Window(tkinter.Toplevel):
    def __init__(self):
        super().__init__()
        self.title('Student List')
        self.geometry('800x800')
        self.minsize(500,500)
        self.maxsize(900,900)
'''
class ObjectData:
    def __init__(self, firstname, lastname, phone_number, student_address, term_period, class_type, class_number, amount_paid, balance):
        # self.name = student_id
        self.age = firstname
        self.age = lastname
        self.age = phone_number
        self.age = student_address
        self.age = term_period
        self.age = class_type
        self.age = class_number
        self.age = amount_paid
        self.age = balance
'''

def fetch_records():
    # global additional_window
    if os.path.exists(filepath):
        # fetch_ids()
        class_type = ask_user_sheet_type()
        additional_window = tkinter.Toplevel()
        additional_window.title(f'{class_type} students list')
        additional_window.geometry('800x800')
        # TODO: make sure file obj is availalable
        # workbook object is created
        wb_obj = openpyxl.load_workbook(filepath)
        
        sheet_obj = wb_obj[class_type]#wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        
        # Loop will print all rows and columns
        for c in range(1, max_col + 1):
            for r in range(1, max_row+1):
                cell_obj = sheet_obj.cell(row = r, column = c)
                ttk.Label(additional_window, text=cell_obj.value).grid(row=r+1, column=c,sticky="news", padx=1, pady=2)
    else:
        tkinter.messagebox.showwarning(title="Error", message=f"Excel file may not exist - Try submitting a record")

def delete_record():
    if os.path.exists(filepath):
        class_type = ask_user_sheet_type()
        fetch_ids(class_type)
        student_Id = row_number[0].get()
        # load excel file
        book = openpyxl.load_workbook(filepath)
    
        # select the sheet
        sheet = book[class_type]
    
        if student_Id in record_ids_dic:
            sheet.delete_rows(record_ids_dic[student_Id], 1)
            book.save(filepath)
            del record_ids_dic[student_Id]
            tkinter.messagebox.showinfo(title='Success', message=f"Successfully deleted record ID: {student_Id}")
        else:
            tkinter.messagebox.showwarning(title="Failure", message=f"Record ID does not exist")
    else:
        tkinter.messagebox.showwarning(title="Error", message=f"Excel file may not exist - Try submitting a record")

def ask_user_sheet_type():
    i = 1
    resp = messagebox.askquestion('Users Response Window', f'What classroom window to display: {class_type_list[i]} sheet?')
    while resp == 'no':
        i+=1
        if i < len(class_type_list):
            resp = messagebox.askquestion('Users Response Window', f'What classroom window to display: {class_type_list[i]} sheet?')
        else:
            messagebox.showinfo('Users Response Window', f'No more classroom windows to display - redirecting to a blank window')
            return 'Sheet'
            
    return class_type_list[i]


def fetch_ids(class_type):
    global record_ids_dic
    record_ids_dic = {}
    if os.path.exists(filepath):
        # resp = ask_user_sheet_type()
        # current_sheet.insert(0, resp)
        book = openpyxl.load_workbook(filepath)

        # select the sheet
        sheet = book[class_type]
        max_row = sheet.max_row

        # Loop will fetch all values from the ID column as key
        # and assign it row number as value to a dictionary.
        for r in range(2, max_row+1):
            cell_obj = sheet.cell(row = r, column = 1)
            record_ids_dic[cell_obj.value] = r

    else:
        tkinter.messagebox.showwarning(title="Error", message=f"Excel file may not exist")
# Not in use currently
def sort_records():
    if os.path.exists(filepath):
        # xlsx = pd.read_ExcelFile(filepath)#df = pd.read_ExcelFile(filepath)
        df = pd.read_excel(filepath, "nusery")
        df.sort_values("ID")
        df.to_excel(filepath, sheet_name="nusery", index=False)
        # df_sorted.to_excel(filepath)
        # df_sorted = df.sort_values('ID')
        # df_sorted.to_excel(filepath)

def delete_view():
    global del_window
    del_window = tkinter.Toplevel()
    del_window.title('Delete Records')
    del_window.geometry('300x300')
    set_form_label(del_window,'Enter ID', 0, 0)
    entry_value = set_form_entry(del_window, 0, 1)
    row_number.append(entry_value)
    ttk.Button(del_window,text='delete', command=delete_record).grid(row=0, column=2, padx=5, pady=10)

# Not currently in use
def close_windows():
    del_window.destroy()

def submit_data():
    accepted = formField_list[9].get()

    if accepted=="Accepted":
        # User info
        # All fields are required except address and phone number.
        firstname = formField_list[0].get()
        lastname = formField_list[1].get()
        student_id = formField_list[2].get()
        # Course info
        amount_paid = formField_list[5].get()
        term_period = formField_list[6].get()
        class_type = formField_list[7].get()
        class_number = formField_list[8].get()

        balance = 0
        total_cost = 650
        if firstname and lastname and student_id and term_period and class_type and class_number and amount_paid:
            amount_paid = formField_list[5].get()
            if int(amount_paid) > 0 and int(amount_paid) <= total_cost:
                balance = total_cost - int(amount_paid)
                # TODO: Optimize formField DS
                student_address = formField_list[3].get()
                phone_number = formField_list[4].get()
                fetch_ids(class_type)

                    # prevent key duplicate
                if student_id in record_ids_dic:
                    tkinter.messagebox.showwarning(title="Error", message=f"Record ID {student_id} does exist, all IDs should be unique.")
                else:    
                    # add new record
                    # student_record = ObjectData(firstname, lastname, phone_number, student_address, term_period, class_type, class_number, int(amount_paid), balance)
                    # record_dic[student_id] = student_record
                    print("---------------Student Data---------------")
                    print(f"First name: {firstname}, Last name: {lastname} \nAmount Paid: le{amount_paid}")
                    print(f"ID: {student_id} \nAddress: {student_address} \nPhone Number: {phone_number}")
                    print(f"Term Period: {term_period} \nClass Type: {class_type} \nLevel: {class_number} \nBalance: le{balance}")
                    # print("Registration status", registration_status)
                    print("------------------------------------------")

                    if not os.path.exists(filepath):
                        workbook = openpyxl.Workbook()
                        # sheet = workbook.active
                        # sheet.append(heading)
                        heading = ["ID", "First Name", "Last Name", "Phone Number", "Address","Term Period", "Class Type", "Level", "Amount Paid", "Balance Due"]

                        workbook.create_sheet('nusery')
                        sheet = workbook[class_type_list[0]]
                        sheet.append(heading)

                        workbook.create_sheet('class')
                        sheet = workbook[class_type_list[1]]
                        sheet.append(heading)
   
                        workbook.create_sheet('JSS')
                        sheet = workbook[class_type_list[2]]
                        sheet.append(heading)

                        workbook.create_sheet('SSS')
                        sheet = workbook[class_type_list[3]]
                        sheet.append(heading)
                        
                        workbook.save(filepath)
                    workbook = openpyxl.load_workbook(filepath)
                    if class_type == class_type_list[1]:
                        sheet = workbook['nusery']
                    elif class_type == class_type_list[2]:
                        sheet = workbook['class']
                    elif class_type == class_type_list[3]:
                        sheet = workbook['JSS']
                    elif class_type == class_type_list[4]:
                        sheet = workbook['SSS']
                    # sheet = workbook.active
                    sheet.append([student_id, firstname, lastname, phone_number, student_address, term_period, class_type, class_number, int(amount_paid), balance])
                    workbook.save(filepath)
                    tkinter.messagebox.showinfo(title="Success", message=f'Record submitted - Thank You!')

            elif int(amount_paid) <= 0 or int(amount_paid) > total_cost:
                tkinter.messagebox.showwarning(title="Error", message="Amount cannot be less than or equal to zero or more than le650.")
                
        else:
            tkinter.messagebox.showwarning(title="Error", message="All fields are required except Address and Phone Number.")
    else:
        tkinter.messagebox.showwarning(title= "Error", message="You have not accepted the terms")

def clear_form():
    # TODO: Optimize the logic in this function
    formField_list[0].delete(0, tkinter.END)
    formField_list[1].delete(0, tkinter.END)
    formField_list[2].delete(0, tkinter.END)
    formField_list[3].delete(0, tkinter.END)
    formField_list[4].delete(0, tkinter.END)
    formField_list[5].delete(0, tkinter.END)
    formField_list[6].delete(0, tkinter.END)
    formField_list[7].delete(0, tkinter.END)
    formField_list[8].delete(0, tkinter.END)


def set_form_label_frame(parentFrame, frameTxt=None, r=0, c=0):
    form_label_frame =tkinter.LabelFrame(parentFrame, text=frameTxt, background='Gray')
    form_label_frame.grid(row= r, column=c)
    return form_label_frame

def set_form_label(labelFrame, labelTxt, r, c):
    form_label = tkinter.Label(labelFrame, text=labelTxt)
    form_label.grid(row=r, column=c)
    return form_label

def set_form_entry(entryFrame, r, c):
    form_entry = tkinter.Entry(entryFrame)
    form_entry.grid(row=r, column=c)
    return form_entry

def set_form_spinbox(spinboxFrame, strt, end, r, c):
    form_spinbox = tkinter.Spinbox(spinboxFrame, from_=strt, to=end)
    form_spinbox.grid(row=r, column=c)
    return form_spinbox

def set_form_combobox(comboboxFrame, class_type_list, r,c):
    form_combobox = ttk.Combobox(comboboxFrame, values=class_type_list)
    form_combobox.grid(row=r, column=c)
    return form_combobox

def create_button(frame, btnType, func, r, c):
     # Button
    button = tkinter.Button(frame, text=btnType, command=func, font=("Helvetica", 10), width=1)
    button.grid(row=r, column=c, sticky="news", padx=5, pady=10)
    return button

# Creating windows Forms
def setup_form(frame):
        # Saving User Info
    student_info_frame = set_form_label_frame(frame, "Student Information")
    set_form_label(student_info_frame,"First Name", 0, 0)
    first_name_entry = set_form_entry(student_info_frame, 0, 1)
    formField_list.append(first_name_entry)

    set_form_label(student_info_frame,"Last Name", 1,0)
    last_name_entry = set_form_entry(student_info_frame, 1,1)
    formField_list.append(last_name_entry)

    set_form_label(student_info_frame,"Student ID", 2, 0)
    student_id_entry = set_form_entry(student_info_frame, 2,1)
    formField_list.append(student_id_entry)

    set_form_label(student_info_frame,"Address", 3,0)
    student_address_entry = set_form_entry(student_info_frame, 3,1)
    formField_list.append(student_address_entry)

    set_form_label(student_info_frame,"Phone Number", 4, 0) 
    phone_number_entry = set_form_entry(student_info_frame, 4,1)
    formField_list.append(phone_number_entry)

    set_form_label(student_info_frame,"Amount Paid", 5, 0)
    amount_paid_entry = set_form_entry(student_info_frame, 5,1)
    formField_list.append(amount_paid_entry)


    for widget in student_info_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    register_status_frame = set_form_label_frame(frame,"Registration Status",1,0)


    set_form_label(register_status_frame, "Term", 0,0)
    term_spinbox = set_form_spinbox(register_status_frame,1,3, 1, 0)
    formField_list.append(term_spinbox)

    set_form_label(register_status_frame, "Class Type", 0,2)
    class_type_combobox = set_form_combobox(register_status_frame, class_type_list ,1,2)
    formField_list.append(class_type_combobox)

    set_form_label(register_status_frame, "Level", 0, 3)
    class_number_spinbox = set_form_spinbox(register_status_frame,1,6, 1, 3)
    formField_list.append(class_number_spinbox)

    for widget in register_status_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    terms_frame = set_form_label_frame(frame, "Terms & Conditions", 2,0)
    terms_frame.rowconfigure(0, weight=1)
    terms_frame.columnconfigure((0,1,2,3), weight=1, uniform='a')

    accept_var = tkinter.StringVar(value="Not Accepted")
    terms_check = tkinter.Checkbutton(terms_frame, text= "I accept the terms and conditions.",
                                    variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
    terms_check.grid(row=0, column=1)
    '''TODO: Change the formField to an object'''
    formField_list.append(accept_var)

    create_button(terms_frame,btnType="Clear Entry", func=clear_form, r=1, c=0)
    create_button(terms_frame,btnType="Submit Data", func=submit_data, r=1, c=2)
    create_button(terms_frame,btnType="List Students Entry", func=fetch_records, r=1, c=1)
    create_button(terms_frame,btnType="Delete Record", func=delete_view, r=1, c=3)


 # TODO: Add display of the student list when inputs are made

def main():
    # code executions
    window = tkinter.Tk()
    window.title("Main Window")
    frame = tkinter.Frame(window)
    frame.pack()
    # fetch_ids()
    setup_form(frame)

    window.mainloop()

if __name__ == "__main__":
    main()